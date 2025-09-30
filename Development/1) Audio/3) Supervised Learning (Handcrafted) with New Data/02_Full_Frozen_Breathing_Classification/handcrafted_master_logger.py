#!/usr/bin/env python3
"""
Handcrafted Features Master Logger
==================================
Creates and manages master Excel log for handcrafted features experiments
"""

import pandas as pd
import numpy as np
from pathlib import Path
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class HandcraftedMasterLogger:
    def __init__(self):
        self.log_file = "Handcrafted_Features_Classification_Log.xlsx"
        self.workbook = None
        
    def create_workbook(self):
        """Create the master Excel workbook with all sheets."""
        
        print("📊 Creating Handcrafted Features Master Log...")
        
        # Create workbook
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets
        self.create_summary_dashboard()
        self.create_individual_models_sheet()
        self.create_ensemble_methods_sheet()
        self.create_temporal_resolution_sheet()
        self.create_feature_analysis_sheet()
        self.create_data_splitting_sheet()
        self.create_error_analysis_sheet()
        
        # Save workbook
        self.workbook.save(self.log_file)
        print(f"✅ Master log created: {self.log_file}")
        
    def create_summary_dashboard(self):
        """Create summary dashboard sheet."""
        
        ws = self.workbook.create_sheet("Summary_Dashboard")
        
        # Headers
        headers = [
            'Experiment_ID', 'Date', 'Experiment_Type', 'Segment_Size', 
            'Method_Type', 'Best_Model', 'Best_Accuracy', 'Improvement', 
            'Status', 'Notes'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ['HCF_0.25s_Individual_Models', '2025-09-22', 'Individual', '0.25s', 'Individual', 'Unknown', 'Unknown', 0.0, 'Pending', 'Handcrafted Features'],
            ['HCF_0.5s_Individual_Models', '2025-09-22', 'Individual', '0.5s', 'Individual', 'Unknown', 'Unknown', 0.0, 'Pending', 'Handcrafted Features'],
            ['HCF_1.0s_Individual_Models', '2025-09-22', 'Individual', '1.0s', 'Individual', 'Unknown', 'Unknown', 0.0, 'Pending', 'Handcrafted Features'],
            ['HCF_0.25s_Ensemble_Models', '2025-09-22', 'Ensemble', '0.25s', 'Ensemble', 'Unknown', 'Unknown', 0.0, 'Pending', 'Handcrafted Features'],
            ['HCF_0.5s_Ensemble_Models', '2025-09-22', 'Ensemble', '0.5s', 'Ensemble', 'Unknown', 'Unknown', 0.0, 'Pending', 'Handcrafted Features'],
            ['HCF_1.0s_Ensemble_Models', '2025-09-22', 'Ensemble', '1.0s', 'Ensemble', 'Unknown', 'Unknown', 0.0, 'Pending', 'Handcrafted Features']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
    def create_individual_models_sheet(self):
        """Create individual models sheet."""
        
        ws = self.workbook.create_sheet("Individual_Models")
        
        # Headers
        headers = [
            'Experiment_ID', 'Date', 'Segment_Size', 'Hop_Size', 
            'Random_Forest_Acc', 'SVM_Acc', 'Logistic_Regression_Acc', 
            'Best_Individual', 'Best_Accuracy', 'Total_Segments', 
            'Breathing_Segments', 'Non_Breathing_Segments', 'Notes'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ['HCF_0.25s_Individual_Models', '2025-09-22', 0.25, 0.25, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Handcrafted Features'],
            ['HCF_0.5s_Individual_Models', '2025-09-22', 0.5, 0.5, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Handcrafted Features'],
            ['HCF_1.0s_Individual_Models', '2025-09-22', 1.0, 1.0, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Handcrafted Features']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
    def create_ensemble_methods_sheet(self):
        """Create ensemble methods sheet."""
        
        ws = self.workbook.create_sheet("Ensemble_Methods")
        
        # Headers
        headers = [
            'Experiment_ID', 'Date', 'Segment_Size', 'Hop_Size', 
            'Individual_Best_Acc', 'Ensemble_Method', 'Ensemble_Config', 
            'Ensemble_Accuracy', 'Improvement', 'Vote_Type', 'Weights', 
            'Models_Used', 'Winner', 'Notes'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ['HCF_0.25s_Ensemble_Models', '2025-09-22', 0.25, 0.25, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 0.0, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Handcrafted Features'],
            ['HCF_0.5s_Ensemble_Models', '2025-09-22', 0.5, 0.5, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 0.0, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Handcrafted Features'],
            ['HCF_1.0s_Ensemble_Models', '2025-09-22', 1.0, 1.0, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 0.0, 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Handcrafted Features']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
    def create_temporal_resolution_sheet(self):
        """Create temporal resolution sheet."""
        
        ws = self.workbook.create_sheet("Temporal_Resolution")
        
        # Headers
        headers = [
            'Experiment_ID', 'Date', 'Segment_Size', 'Hop_Size', 
            'Overlap_Percentage', 'Labeling_Method', 'Accuracy', 
            'Precision', 'Recall', 'F1_Score', 'Total_Test_Segments', 
            'Visual_Mathematical_Match', 'Notes'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ['HCF_0.25s_Individual_Models', '2025-09-22', 0.25, 0.25, 0, 'Center-Point', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Yes', 'Handcrafted Features'],
            ['HCF_0.5s_Individual_Models', '2025-09-22', 0.5, 0.5, 0, 'Center-Point', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Yes', 'Handcrafted Features'],
            ['HCF_1.0s_Individual_Models', '2025-09-22', 1.0, 1.0, 0, 'Center-Point', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Yes', 'Handcrafted Features']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
    def create_feature_analysis_sheet(self):
        """Create feature analysis sheet."""
        
        ws = self.workbook.create_sheet("Feature_Analysis")
        
        # Headers
        headers = [
            'Experiment_ID', 'Date', 'Feature_Type', 'Feature_Count', 
            'Handcrafted_Features', 'OPERA_CT_Features', 'Combined_Features', 
            'Best_Accuracy', 'Feature_Importance_Available', 'Notes'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ['HCF_All_Models', '2025-09-22', 'Handcrafted', '~50-100', 'Yes', 'No', 'No', 'Unknown', 'Yes', 'Traditional signal processing features']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
    def create_data_splitting_sheet(self):
        """Create data splitting sheet."""
        
        ws = self.workbook.create_sheet("Data_Splitting")
        
        # Headers
        headers = [
            'Experiment_ID', 'Date', 'Split_Method', 'Train_Size', 
            'Test_Size', 'Patient_Based', 'Segment_Based', 'Stratified', 
            'Accuracy', 'Generalization_Score', 'Notes'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ['HCF_All_Models', '2025-09-22', 'Random', 'Unknown', 'Unknown', 'No', 'Yes', 'Yes', 'Unknown', 'Unknown', 'Handcrafted Features']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
    def create_error_analysis_sheet(self):
        """Create error analysis sheet."""
        
        ws = self.workbook.create_sheet("Error_Analysis")
        
        # Headers
        headers = [
            'Experiment_ID', 'Date', 'Segment_Size', 'Method', 
            'False_Positives', 'False_Negatives', 'Common_Errors', 
            'Problematic_Files', 'Error_Patterns', 'Recommendations'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add sample data
        sample_data = [
            ['HCF_All_Models', '2025-09-22', 'All', 'All', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Unknown', 'Handcrafted Features']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
    def update_results(self, experiment_id, results_data):
        """Update results in the appropriate sheet."""
        
        print(f"📝 Updating results for {experiment_id}...")
        
        # This would be implemented to update specific sheets with actual results
        # For now, just save the workbook
        self.workbook.save(self.log_file)
        print(f"✅ Results updated in {self.log_file}")

def main():
    """Main execution function."""
    
    print("📊 HANDCRAFTED FEATURES MASTER LOGGER")
    print("=" * 50)
    
    logger = HandcraftedMasterLogger()
    logger.create_workbook()
    
    print("\n✅ Master Excel log created successfully!")
    print("📋 Sheets created:")
    print("   • Summary_Dashboard")
    print("   • Individual_Models")
    print("   • Ensemble_Methods")
    print("   • Temporal_Resolution")
    print("   • Feature_Analysis")
    print("   • Data_Splitting")
    print("   • Error_Analysis")
    print("\n🎯 Ready for handcrafted features experiments!")

if __name__ == "__main__":
    main()

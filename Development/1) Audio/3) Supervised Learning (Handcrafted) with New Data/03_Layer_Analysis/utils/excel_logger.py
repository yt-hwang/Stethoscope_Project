#!/usr/bin/env python3
"""
Excel Logger - Global workbook management with structured sheets
Uses openpyxl for append-only logging without data assumptions
"""

import openpyxl
from openpyxl import Workbook
from pathlib import Path
import os
from datetime import datetime

class ExcelLogger:
    """Manages global experiment_log.xlsx with structured sheets."""
    
    def __init__(self, base_results_dir="results"):
        self.base_results_dir = Path(base_results_dir)
        self.excel_path = self.base_results_dir / "experiment_log.xlsx"
        self.base_results_dir.mkdir(parents=True, exist_ok=True)
        
        # Define required sheets
        self.required_sheets = ["runs", "features", "train", "eval", "cluster", "errors"]
        
        # Initialize workbook if it doesn't exist
        self._init_workbook()
    
    def _init_workbook(self):
        """Initialize workbook with required sheets if it doesn't exist."""
        if not self.excel_path.exists():
            wb = Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Create required sheets
            for sheet_name in self.required_sheets:
                wb.create_sheet(sheet_name)
            
            wb.save(self.excel_path)
            print(f"✅ Created experiment_log.xlsx with sheets: {', '.join(self.required_sheets)}")
        else:
            # Check if all required sheets exist
            wb = openpyxl.load_workbook(self.excel_path)
            missing_sheets = []
            
            for sheet_name in self.required_sheets:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                    missing_sheets.append(sheet_name)
            
            if missing_sheets:
                wb.save(self.excel_path)
                print(f"✅ Added missing sheets: {', '.join(missing_sheets)}")
    
    def append_row(self, sheet_name, data_dict):
        """
        Append a row to specified sheet.
        Auto-creates headers if sheet is empty.
        """
        if sheet_name not in self.required_sheets:
            raise ValueError(f"Sheet '{sheet_name}' not in required sheets: {self.required_sheets}")
        
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        # Add timestamp if not present
        if 'timestamp' not in data_dict:
            data_dict['timestamp'] = datetime.now().isoformat()
        
        # If sheet is empty, create headers
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            headers = list(data_dict.keys())
            for col, header in enumerate(headers, 1):
                ws.cell(1, col, header)
            
            # Add data row
            for col, value in enumerate(data_dict.values(), 1):
                ws.cell(2, col, value)
        else:
            # Get existing headers
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(1, col).value
                if header:
                    headers.append(header)
            
            # Add any new headers
            all_keys = set(headers + list(data_dict.keys()))
            if len(all_keys) > len(headers):
                for col, key in enumerate(sorted(all_keys), 1):
                    ws.cell(1, col, key)
                headers = sorted(all_keys)
            
            # Add data row
            new_row = ws.max_row + 1
            for col, header in enumerate(headers, 1):
                value = data_dict.get(header, "")
                ws.cell(new_row, col, value)
        
        wb.save(self.excel_path)
        return ws.max_row - 1  # Return row number (excluding header)
    
    def get_sheet_data(self, sheet_name):
        """Get all data from a sheet as list of dictionaries."""
        if sheet_name not in self.required_sheets:
            raise ValueError(f"Sheet '{sheet_name}' not in required sheets: {self.required_sheets}")
        
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        if ws.max_row <= 1:
            return []  # Empty sheet
        
        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header:
                headers.append(header)
        
        # Get data rows
        data = []
        for row in range(2, ws.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                value = ws.cell(row, col).value
                row_data[header] = value
            data.append(row_data)
        
        return data
    
    def update_run_row(self, sheet_name, run_id, update_dict):
        """Update an existing row identified by run_id."""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb[sheet_name]
        
        # Find the row with matching run_id
        run_id_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(1, col).value == 'run_id':
                run_id_col = col
                break
        
        if run_id_col is None:
            raise ValueError(f"No 'run_id' column found in sheet '{sheet_name}'")
        
        # Find matching row
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, run_id_col).value == run_id:
                target_row = row
                break
        
        if target_row is None:
            # Row doesn't exist, append new row
            return self.append_row(sheet_name, {'run_id': run_id, **update_dict})
        
        # Update existing row
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header:
                headers.append(header)
        
        # Update values
        for key, value in update_dict.items():
            if key in headers:
                col = headers.index(key) + 1
                ws.cell(target_row, col, value)
            else:
                # Add new column
                new_col = len(headers) + 1
                ws.cell(1, new_col, key)
                ws.cell(target_row, new_col, value)
                headers.append(key)
        
        wb.save(self.excel_path)
        return target_row - 1  # Return row number (excluding header)

# Convenience functions
def append_to_excel(sheet_name, data_dict, results_dir="results"):
    """Append row to Excel sheet (convenience function)."""
    logger = ExcelLogger(results_dir)
    return logger.append_row(sheet_name, data_dict)

def update_excel_row(sheet_name, run_id, update_dict, results_dir="results"):
    """Update Excel row (convenience function)."""
    logger = ExcelLogger(results_dir)
    return logger.update_run_row(sheet_name, run_id, update_dict)

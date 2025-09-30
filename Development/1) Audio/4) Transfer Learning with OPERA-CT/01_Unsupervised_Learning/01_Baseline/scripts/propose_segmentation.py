#!/usr/bin/env python3
"""
Segmentation Plan Script - Propose segmentation based on audited durations
No processing yet - only statistical analysis and planning
"""

import argparse
import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from pathlib import Path
from datetime import datetime

# Add utils to path
sys.path.append(str(Path(__file__).parent.parent / "utils"))
from excel_logger import ExcelLogger

def analyze_durations(df):
    """Analyze duration statistics from data audit."""
    durations = df['duration_sec'].dropna()
    
    if len(durations) == 0:
        return None
    
    stats = {
        'count': len(durations),
        'mean': float(np.mean(durations)),
        'median': float(np.median(durations)),
        'std': float(np.std(durations)),
        'min': float(np.min(durations)),
        'max': float(np.max(durations)),
        'q25': float(np.percentile(durations, 25)),
        'q75': float(np.percentile(durations, 75))
    }
    
    return stats

def propose_segment_candidates(duration_stats):
    """Propose 3 segmentation candidates based on duration statistics."""
    median_dur = duration_stats['median']
    
    # Calculate candidates as percentages of median duration
    candidates = {
        'seg_short': {
            'target_percent': 15,  # 10-20% range, use 15%
            'duration_sec': median_dur * 0.15,
            'description': 'Short segments (15% of median duration)'
        },
        'seg_mid': {
            'target_percent': 25,  # 15-30% range, use 25% 
            'duration_sec': median_dur * 0.25,
            'description': 'Medium segments (25% of median duration)'
        },
        'seg_long': {
            'target_percent': 35,  # 25-40% range, use 35%
            'duration_sec': median_dur * 0.35,
            'description': 'Long segments (35% of median duration)'
        }
    }
    
    return candidates

def estimate_segment_counts(df, candidates):
    """Estimate segment counts for each candidate and overlap setting."""
    results = []
    
    for candidate_name, candidate_info in candidates.items():
        seg_duration = candidate_info['duration_sec']
        
        for overlap_percent in [0, 50]:  # No overlap and 50% overlap
            overlap_sec = seg_duration * (overlap_percent / 100)
            step_sec = seg_duration - overlap_sec
            
            total_segments = 0
            file_estimates = []
            
            for _, row in df.iterrows():
                if pd.isna(row['duration_sec']):
                    continue
                    
                file_duration = row['duration_sec']
                
                if file_duration < seg_duration:
                    # File too short - no segments
                    n_segments = 0
                else:
                    # Calculate number of segments
                    n_segments = int((file_duration - seg_duration) / step_sec) + 1
                
                total_segments += n_segments
                
                file_estimates.append({
                    'filename': row['filename'],
                    'file_duration': file_duration,
                    'n_segments': n_segments
                })
            
            # Determine if this is a good configuration
            recommended = 150 <= total_segments <= 500
            
            result = {
                'candidate': candidate_name,
                'seg_duration': seg_duration,
                'overlap_percent': overlap_percent,
                'overlap_sec': overlap_sec,
                'step_sec': step_sec,
                'total_segments': total_segments,
                'recommended': recommended,
                'description': f"{candidate_info['description']} - {overlap_percent}% overlap"
            }
            
            results.append((result, file_estimates))
    
    return results

def create_segmentation_plan_excel(run_id, candidates, segment_estimates, duration_stats):
    """Create Excel file with segmentation plan."""
    run_dir = Path("results/experiments") / run_id
    artifacts_dir = run_dir / "artifacts"
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = artifacts_dir / "segmentation_plan.xlsx"
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Sheet 1: Plan overview
    ws_plan = wb.create_sheet("plan")
    
    # Headers for plan sheet
    plan_headers = [
        'candidate', 'seg_duration', 'overlap_percent', 'overlap_sec', 
        'step_sec', 'total_segments', 'recommended', 'description'
    ]
    
    for col, header in enumerate(plan_headers, 1):
        ws_plan.cell(1, col, header)
    
    # Add plan data
    for row_idx, (result, _) in enumerate(segment_estimates, 2):
        for col, header in enumerate(plan_headers, 1):
            ws_plan.cell(row_idx, col, result[header])
    
    # Sheet 2: By-file projections
    ws_files = wb.create_sheet("by_file_projection")
    
    # Headers for file projection sheet
    file_headers = ['candidate', 'overlap_percent', 'filename', 'file_duration', 'n_segments']
    
    for col, header in enumerate(file_headers, 1):
        ws_files.cell(1, col, header)
    
    # Add file projection data
    row_idx = 2
    for result, file_estimates in segment_estimates:
        for file_est in file_estimates:
            ws_files.cell(row_idx, 1, result['candidate'])
            ws_files.cell(row_idx, 2, result['overlap_percent'])
            ws_files.cell(row_idx, 3, file_est['filename'])
            ws_files.cell(row_idx, 4, file_est['file_duration'])
            ws_files.cell(row_idx, 5, file_est['n_segments'])
            row_idx += 1
    
    # Sheet 3: Duration statistics
    ws_stats = wb.create_sheet("duration_stats")
    
    # Add duration statistics
    stats_data = [
        ['Statistic', 'Value'],
        ['Count', duration_stats['count']],
        ['Mean', duration_stats['mean']],
        ['Median', duration_stats['median']],
        ['Std Dev', duration_stats['std']],
        ['Min', duration_stats['min']],
        ['Max', duration_stats['max']],
        ['Q25', duration_stats['q25']],
        ['Q75', duration_stats['q75']]
    ]
    
    for row_idx, (stat, value) in enumerate(stats_data, 1):
        ws_stats.cell(row_idx, 1, stat)
        ws_stats.cell(row_idx, 2, value)
    
    # Save workbook
    wb.save(excel_path)
    
    return excel_path

def main():
    parser = argparse.ArgumentParser(description="Propose segmentation plan based on audited durations")
    parser.add_argument("--run_id", required=True, help="Run ID")
    parser.add_argument("--root", required=True, help="Data root path (for reference)")
    
    args = parser.parse_args()
    
    print(f"Segmentation planning for run: {args.run_id}")
    
    # Load data audit results
    run_dir = Path("results/experiments") / args.run_id
    audit_csv_path = run_dir / "artifacts" / "data_audit.csv"
    
    if not audit_csv_path.exists():
        print(f"Error: Data audit CSV not found: {audit_csv_path}")
        print("Please run data_audit.py first.")
        sys.exit(1)
    
    print("Step 1: Loading data audit results...")
    df = pd.read_csv(audit_csv_path)
    print(f"Loaded {len(df)} file records")
    
    print("Step 2: Analyzing duration statistics...")
    duration_stats = analyze_durations(df)
    
    if duration_stats is None:
        print("Error: No valid duration data found")
        sys.exit(1)
    
    print(f"Duration stats: median={duration_stats['median']:.1f}s, range={duration_stats['min']:.1f}-{duration_stats['max']:.1f}s")
    
    print("Step 3: Proposing segmentation candidates...")
    candidates = propose_segment_candidates(duration_stats)
    
    for name, info in candidates.items():
        print(f"  {name}: {info['duration_sec']:.1f}s ({info['description']})")
    
    print("Step 4: Estimating segment counts...")
    segment_estimates = estimate_segment_counts(df, candidates)
    
    print("Step 5: Creating segmentation plan Excel...")
    excel_path = create_segmentation_plan_excel(args.run_id, candidates, segment_estimates, duration_stats)
    print(f"Segmentation plan saved: {excel_path}")
    
    print("Step 6: Logging to main Excel...")
    # Log to main experiment Excel
    excel_logger = ExcelLogger()
    
    excel_data = {
        'run_id': args.run_id,
        'segmentation_plan_path': str(excel_path),
        'median_duration': duration_stats['median'],
        'n_candidates': len(candidates),
        'recommended_configs': sum(1 for result, _ in segment_estimates if result['recommended'])
    }
    
    try:
        row_num = excel_logger.append_row('runs', excel_data)
        print(f"Main Excel updated: row {row_num}")
    except Exception as e:
        print(f"Warning: Excel logging failed: {e}")
    
    print("RESULTS SUMMARY:")
    print(f"  Median duration: {duration_stats['median']:.1f}s")
    print(f"  Segmentation candidates: {len(candidates)}")
    
    # Show recommendations
    recommended_count = 0
    for result, _ in segment_estimates:
        if result['recommended']:
            recommended_count += 1
            print(f"  RECOMMENDED: {result['candidate']} {result['overlap_percent']}% overlap -> {result['total_segments']} segments")
    
    if recommended_count == 0:
        print("  No configurations meet target range (150-500 segments)")
    
    print(f"Excel plan: {excel_path}")
    print("STOP - Segmentation planning complete.")

if __name__ == "__main__":
    main()

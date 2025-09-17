#!/usr/bin/env python3
"""
Log D1_new cycle results to Excel tracking system
"""

import pandas as pd
import os
import json
from openpyxl import load_workbook
import argparse

# Configuration
EXCEL_PATH = '/Users/yunhwang/Desktop/Stethoscope_Project/Experiment_Tracking_System_Final.xlsx'
OUTPUT_ROOT = '/Users/yunhwang/Desktop/Stethoscope_Project/Development/cycles/D1_new_Seg_PeakNormalize_Bandpass/outputs'
CYCLE_ID = 'D1_new'
PHASE = 'D-Series_New'
SEGMENTATION = 'Seg_10sec'
PREPROCESSING = 'peak_normalize_bandpass_100_2000'
REPRESENTATIONS = 'raw_waveform_stats, logmel_mean, mfcc_mean'
CLUSTERING_ALGOS = 'kmeans(k=3,4,5), hdbscan(min_cluster_size=3)'
NOTES = 'Better combination: PeakNormalize + Bandpass (amplitude + frequency)'

def append_cycle_summary(df_summary):
    """Append cycle summary to the CycleSummary sheet."""
    new_row = {
        'Cycle_ID': CYCLE_ID,
        'Phase': PHASE,
        'Segmentation': SEGMENTATION,
        'Preprocessing': PREPROCESSING,
        'Representations': REPRESENTATIONS,
        'Clustering_Algos': CLUSTERING_ALGOS,
        'Output_Root': OUTPUT_ROOT,
        'Notes': NOTES
    }
    return pd.concat([df_summary, pd.DataFrame([new_row])], ignore_index=True)

def append_metrics_by_run(df_metrics, metrics_data, representation):
    """Append detailed metrics to the MetricsByRun sheet."""
    rows_to_add = []
    
    for algo_key, data in metrics_data.items():
        algorithm = data['algorithm']
        params = {}
        if algorithm == 'kmeans':
            params = {'k': data['k']}
        elif algorithm == 'hdbscan':
            params = {'min_cluster_size': data['min_cluster_size']}

        features_csv_path = os.path.join(OUTPUT_ROOT, 'features', f'features_{representation}.csv')
        metrics_json_path = os.path.join(OUTPUT_ROOT, 'clustering', f'clustering_{representation}_metrics.json')
        
        # Construct UMAP PNG path
        if algorithm == 'kmeans':
            umap_png_filename = f"umap_{representation}_kmeans_k{data['k']}.png"
        elif algorithm == 'hdbscan':
            umap_png_filename = f"umap_{representation}_hdbscan_min_cluster_size{data['min_cluster_size']}.png"
        else:
            umap_png_filename = ""

        umap_png_path = os.path.join(OUTPUT_ROOT, 'visualizations', umap_png_filename)

        rows_to_add.append({
            'Cycle_ID': CYCLE_ID,
            'Representation': representation,
            'Clustering': algorithm,
            'Params': json.dumps(params),
            'Silhouette': data.get('silhouette'),
            'CalinskiHarabasz': data.get('calinski_harabasz'),
            'DaviesBouldin': data.get('davies_bouldin'),
            'NumClusters': data.get('num_clusters'),
            'NumNoise': data.get('num_noise_points'),
            'Features_CSV_Path': features_csv_path,
            'Metrics_JSON_Path': metrics_json_path,
            'UMAP_PNG_Path': umap_png_path,
            'Notes': NOTES
        })
    
    return pd.concat([df_metrics, pd.DataFrame(rows_to_add)], ignore_index=True)

def main():
    parser = argparse.ArgumentParser(description="Log D1_new cycle results to Excel.")
    parser.add_argument('--output_dir', type=str, default=OUTPUT_ROOT,
                        help='Path to the output directory for this cycle.')
    
    args = parser.parse_args()

    print("="*60)
    print("D1_new: Logging to Excel")
    print("="*60)

    # Load existing workbook or create if not exists
    try:
        book = load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        print(f"Creating new Excel file at {EXCEL_PATH}")
        book = pd.ExcelWriter(EXCEL_PATH, engine='openpyxl').book
        book.create_sheet("CycleSummary")
        book.create_sheet("MetricsByRun")
        book.create_sheet("Artifacts")
        if 'Sheet' in book.sheetnames:
            del book['Sheet']

    # Read existing sheets
    df_summary = pd.read_excel(EXCEL_PATH, sheet_name='CycleSummary') if 'CycleSummary' in book.sheetnames else pd.DataFrame()
    df_metrics = pd.read_excel(EXCEL_PATH, sheet_name='MetricsByRun') if 'MetricsByRun' in book.sheetnames else pd.DataFrame()

    # Append cycle summary
    df_summary = append_cycle_summary(df_summary)
    print("✓ Added cycle summary")

    # Append metrics for each representation
    representations = ['raw_waveform_stats', 'logmel_mean', 'mfcc_mean']
    for rep in representations:
        metrics_json_path = os.path.join(args.output_dir, 'clustering', f'clustering_{rep}_metrics.json')
        if os.path.exists(metrics_json_path):
            with open(metrics_json_path, 'r') as f:
                metrics_data = json.load(f)
            df_metrics = append_metrics_by_run(df_metrics, metrics_data, rep)
            print(f"✓ Added metrics for {rep}")
        else:
            print(f"Warning: Metrics JSON not found for {rep} at {metrics_json_path}")

    # Write updated DataFrames back to Excel
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='CycleSummary', index=False)
        df_metrics.to_excel(writer, sheet_name='MetricsByRun', index=False)
        
        # Preserve other sheets
        for sheet_name in book.sheetnames:
            if sheet_name not in ['CycleSummary', 'MetricsByRun']:
                book[sheet_name].title = sheet_name
                writer.book.append(book[sheet_name])

    print(f"\n✓ Successfully logged Cycle {CYCLE_ID} results to {EXCEL_PATH}")
    print("="*60)

if __name__ == "__main__":
    main()
